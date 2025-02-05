import os
from expense import Expense
import calendar
import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.drawing.image import Image
import plotly.express as px
import plotly.io as pio
from io import BytesIO
import matplotlib.pyplot as plt
import keyboard



def main():
    print(f"Running Expense Tracker!")
    expense_file_path = os.path.join(os.path.dirname(__file__), "expenses.xlsx")
    budget = 2000

    # Get user to input expense
    expense = get_user_expense()

    # Write it to Excel file
    for expense in expense:
        save_expense_to_excel(expense, expense_file_path)

    # Read file and summarize expenses
    summarize_expense(expense_file_path, budget)

    # Generate and save visualizations
    create_visualizations(expense_file_path)

def get_user_expense():
    print("Getting user expenses!")

    expense_categories = ["Food", "Home", "Work", "Fun", "Misc"]
    expenses = []

    while True:
        print("\nReturn '\\' at any point to add to excel sheet")

        # Get expense name
        expense_name = input("Enter expense name: ")
        if expense_name.lower() == "\\":
            print("Exiting input loop.")
            break

        # Get expense amount
        try:
            expense_amount = input("Enter expense amount: ")
            if expense_amount.lower() == "\\":
                print("Exiting input loop.")
                break
            expense_amount = float(expense_amount)
        except ValueError:
            print("Invalid amount. Please enter a numeric value.")
            continue

        # Select category
        while True:
            print("Select a category: ")
            for i, category_name in enumerate(expense_categories):
                print(f"{i + 1}. {category_name}")

            value_range = f"[1 - {len(expense_categories)}]"
            selected_index = input(f"Enter a category number {value_range}: ")

            if selected_index.lower() == "\\":
                print("Exiting input loop.")
                return expenses

            try:
                selected_index = int(selected_index) - 1
                if selected_index in range(len(expense_categories)):
                    selected_category = expense_categories[selected_index]
                    new_expense = Expense(
                        name=expense_name, category=selected_category, amount=expense_amount
                    )
                    expenses.append(new_expense)
                    print("Expense added successfully!")
                    break
                else:
                    print("Invalid category, please try again!")
            except ValueError:
                print("Invalid input. Please enter a number.")

    return expenses


def save_expense_to_excel(expense: Expense, expense_file_path):
    print(f"Saving expense to Excel file: {expense} to {expense_file_path}")

    if not os.path.exists(expense_file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expenses"
        sheet.append(["Date", "Name", "Amount", "Category"])
    else:
        workbook = load_workbook(expense_file_path)
        sheet = workbook.active

    # Get the current date
    now = datetime.datetime.now().strftime("%Y-%m-%d")

    # Add the new expense to the Excel file
    sheet.append([now, expense.name, expense.amount, expense.category])
    workbook.save(expense_file_path)
    print(f"Expense saved successfully to {expense_file_path}")

def summarize_expense(expense_file_path, budget):
    print(f"Summarizing expenses from Excel file!")

    # Load the Excel file into a DataFrame
    try:
        df = pd.read_excel(expense_file_path, sheet_name="Expenses")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Check the columns and data types
    print("Columns in the Excel file:", df.columns)

    # Ensure the columns are correctly named
    expected_columns = ["Date", "Name", "Amount", "Category"]
    if not all(col in df.columns for col in expected_columns):
        print(f"Unexpected columns found. Expected columns: {expected_columns}")
        return

    # Convert 'Amount' column to numeric, handling any conversion errors
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")

    # Drop rows with invalid 'Amount' values (e.g., non-numeric)
    df.dropna(subset=["Amount"], inplace=True)

    # Calculate expenses by category
    amount_by_category = df.groupby("Category")["Amount"].sum().to_dict()

    # Display expenses by category
    print("Expenses by Category:")
    for key, amount in amount_by_category.items():
        print(f"  {key}: ${amount:.2f}")

    # Calculate total spent and remaining budget
    total_spent = df["Amount"].sum()
    print(f"You've spent ${total_spent:.2f}")

    remaining_budget = budget - total_spent
    print(f"Budget remaining: ${remaining_budget:.2f}")

    # Calculate remaining days in the current month
    now = datetime.datetime.now()
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    remaining_days = days_in_month - now.day
    print("Remaining days in the current month:", remaining_days)

    # Calculate daily budget
    daily_budget = remaining_budget / remaining_days if remaining_days > 0 else 0
    print(f"Budget per day: ${daily_budget:.2f}")
    print(f"Summarizing expenses from Excel file!")
    expenses = []

    # Load the Excel file and read data
    workbook = load_workbook(expense_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        date, expense_name, expense_amount, expense_category = row
        line_expense = Expense(
            name=expense_name, amount=float(expense_amount), category=expense_category
        )
        expenses.append(line_expense)

    # Calculate expenses by category
    amount_by_category = {}
    for expense in expenses:
        key = expense.category
        if key in amount_by_category:
            amount_by_category[key] += expense.amount
        else:
            amount_by_category[key] = expense.amount

    print("Expenses by Category: ")
    for key, amount in amount_by_category.items():
        print(f"  {key}: ${amount:.2f}")

    total_spent = sum([x.amount for x in expenses])
    print(f"You've spent ${total_spent:.2f}")

    remaining_budget = budget - total_spent
    print(f"Budget remaining ${remaining_budget:.2f}")

    now = datetime.datetime.now()
    days_in_month = calendar.monthrange(now.year, now.month)[1]
    remaining_days = days_in_month - now.day
    print("Remaining days in the current month: ", remaining_days)

    daily_budget = remaining_budget / remaining_days if remaining_days > 0 else 0
    print(f"Budget per day ${daily_budget:.2f}")

def create_visualizations(expense_file_path):
    print("Generating visualizations directly in Excel...")

    # Load data from the Excel file into a DataFrame
    df = pd.read_excel(expense_file_path, sheet_name="Expenses")

    # Get the current month and year for the titles
    now = datetime.datetime.now()
    month_year = now.strftime("%B %Y")

    # Remove any existing images (old graphs) from the sheet
    workbook = load_workbook(expense_file_path)
    sheet = workbook.active
    if hasattr(sheet, "_images"):
        sheet._images.clear()

    # Generate Pie Chart: Expenses by Category
    pie_title = f"Spendings by Category for {month_year}"
    fig_pie, ax_pie = plt.subplots()
    category_data = df.groupby("Category")["Amount"].sum()
    ax_pie.pie(category_data, labels=category_data.index, autopct='%1.1f%%', startangle=90)
    ax_pie.set_title(pie_title)
    ax_pie.axis('equal')  # Equal aspect ratio ensures the pie chart is circular

    # Save the pie chart to an in-memory image
    pie_image_stream = BytesIO()
    fig_pie.savefig(pie_image_stream, format='png')
    pie_image_stream.seek(0)
    plt.close(fig_pie)

    # Insert the Pie Chart into Excel
    pie_image = Image(pie_image_stream)
    pie_image.width, pie_image.height = 600, 400
    sheet.add_image(pie_image, "F2")

    # Generate Stacked Histogram: Daily Spending by Category
    line_title = f"Spendings for {month_year} by Category"
    fig_bar, ax_bar = plt.subplots(figsize=(10, 6))
    df['Date'] = pd.to_datetime(df['Date'])
    df['Day'] = df['Date'].dt.day  # Extract day from the date

    # Pivot the data to get daily spending by category
    daily_spending = df.pivot_table(index='Day', columns='Category', values='Amount', aggfunc='sum').fillna(0)

    # Create a stacked bar chart
    categories = daily_spending.columns
    bottom = None
    for category in categories:
        if bottom is None:
            ax_bar.bar(daily_spending.index, daily_spending[category], label=category)
            bottom = daily_spending[category]
        else:
            ax_bar.bar(daily_spending.index, daily_spending[category], bottom=bottom, label=category)
            bottom += daily_spending[category]

    # Customize the bar chart
    ax_bar.set_title(line_title)
    ax_bar.set_xlabel("Day of the Month")
    ax_bar.set_ylabel("Amount Spent ($)")
    ax_bar.set_xticks(daily_spending.index)  # Ensure the x-axis shows only whole numbers (days)
    ax_bar.set_xticklabels(daily_spending.index)
    ax_bar.legend(title="Category")
    ax_bar.grid(axis='y')

    # Save the stacked bar chart to an in-memory image
    bar_image_stream = BytesIO()
    fig_bar.savefig(bar_image_stream, format='png')
    bar_image_stream.seek(0)
    plt.close(fig_bar)

    # Insert the Stacked Histogram into Excel
    bar_image = Image(bar_image_stream)
    bar_image.width, bar_image.height = 600, 400
    sheet.add_image(bar_image, "F20")

    # Save the workbook with the embedded images
    workbook.save(expense_file_path)
    print("Visualizations updated and embedded into the Excel file successfully.")
    
if __name__ == "__main__":
    main()
